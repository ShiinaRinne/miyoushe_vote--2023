import time
import openpyxl
import aiohttp
import asyncio
from typing import Dict, List

from .api import *
from .model import *


async def sem_coro(sem, coro):
    async with sem:
        return await coro


async def save(headers, max_gather: int = 4):
    temp = openpyxl.Workbook()
    temp.remove(temp["Sheet"])
    set_headers(headers)

    all_author_list: Dict[str, List[AuthorInfo]] = {}
    sem = asyncio.Semaphore(max_gather)
    async with aiohttp.ClientSession() as session:
        for t in AuthorType:
            current_sheet = temp.create_sheet(t.value)
            current_sheet.append(["Nickname", "Vote Num Gap", "Timestamp"])

            total = await get_total_users_at_type(t, session)

            tasks = [get_users_at_page(page, t, session) for page in range(1, total // 15 + 1)]
            pages: List[FullAuthorList] = await asyncio.gather(*(sem_coro(sem, task) for task in tasks))

            author_info_list: List[AuthorInfo] = []
            for page in pages:
                author_info_list.extend(page.full_author_list)
            
            all_author_list[t.value] = author_info_list


        for k, v in all_author_list.items():
            current_sheet = temp[k]
            print(f"正在获取{k:15}分区的投票信息")

            tasks = [get_votes_of_user(user.aid, session) for user in v]
            vote_list: List[VotesInfo] = await asyncio.gather(*(sem_coro(sem, task) for task in tasks))

            for user, vote_info in zip(v, vote_list):
                current_sheet.append([user.nickname, vote_info.vote_num_gap, time.time()])

    temp.save("temp.xlsx")


def calculate(file_name):
    temp = openpyxl.load_workbook("temp.xlsx")
    vote_wb = openpyxl.Workbook()
    vote_wb.remove(vote_wb["Sheet"])

    for sheet in temp.worksheets:
        vote_wb.create_sheet(sheet.title)
        vote_wb[sheet.title].append(["Nickname", "total", "Gap"])

        user_info = [(row[0].value, row[1].value, row[2].value) 
                     for row in sheet.iter_rows(min_row=2)][::-1] # nickname, gap, timestamp
        
        users = [] # nickname, total, gap
        for i in range(len(user_info)):
            user_info[i] = (
                user_info[i][0],
                0 if i == 0 else user_info[i - 1][2] + user_info[i - 1][1],
                user_info[i][1],
            )

            users.append(user_info[i])

        for i in users[::-1]:
            vote_wb[sheet.title].append(i)

    vote_wb.save(f"{file_name}.xlsx")
