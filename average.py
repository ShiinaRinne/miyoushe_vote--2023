import openpyxl

days = ["1221", "1222", "1223", "1225", "1227", "1228", "1229", "1230"]

all_data = {
    "AuthorDrawing": {},
    "AuthorStrategy": {},
    "AuthorCoser": {},
    "AuthorOther": {},
}

for day in days:
    temp = openpyxl.load_workbook(f"vote-{day}.xlsx")

    for sheet in temp.worksheets:
        sheet_data = list(
            sheet.iter_rows(min_row=2, values_only=True)
        )  # nickname, total, gap
        for row in sheet_data:
            if all_data[sheet.title].get(row[0], None) is None:
                all_data[sheet.title][row[0]] = {}
            all_data[sheet.title][row[0]][day] = row[1]


wb = openpyxl.Workbook()
wb.remove(wb["Sheet"])
for k, v in all_data.items():
    wb.create_sheet(k)
    wb[k].append(["Nickname"] + days + ["Average"])
    for nickname, data in v.items():
        wb[k].append(
            [nickname]
            + [data.get(day, 0) for day in days]
            + [(data.get(days[-1], 0) - data.get(days[0], 0)) / 10]
        )

wb.save("all_data.xlsx")
